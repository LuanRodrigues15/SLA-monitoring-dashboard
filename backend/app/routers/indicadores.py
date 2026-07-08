import io

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse

from app.core.trino_client import TrinoClient
from app.deps import get_trino, get_current_user
from app.models.auth import UserOut
from app.models.kpi import KpiDetail, KpiHistoricoItem
from app.services.kpi_calculator import get_kpi_detail, get_kpi_historico

router = APIRouter(prefix="/api/indicadores", tags=["indicadores"])


@router.get("/{cod}", response_model=KpiDetail)
def indicador_detalhe(
    cod: str,
    competencia: str | None = Query(default=None),
    tabela: str = Query(default="kpi_agg_test", description="Tabela de leitura: kpi_agg_test (padrão) ou kpi_agg"),
    trino: TrinoClient = Depends(get_trino),
    _user: UserOut = Depends(get_current_user),
):
    return get_kpi_detail(trino, cod, competencia, tabela)


@router.get("/{cod}/historico", response_model=list[KpiHistoricoItem])
def indicador_historico(
    cod: str,
    meses: int = Query(default=12, ge=1, le=60),
    trino: TrinoClient = Depends(get_trino),
    _user: UserOut = Depends(get_current_user),
):
    return get_kpi_historico(trino, cod, meses)


@router.get("/{cod}/excel")
def indicador_excel(
    cod: str,
    competencia: str | None = Query(default=None),
    trino: TrinoClient = Depends(get_trino),
    _user: UserOut = Depends(get_current_user),
):
    detail = get_kpi_detail(trino, cod, competencia)
    comp = competencia or ""
    excel_bytes = _build_detail_excel(detail, comp)
    filename = f"KPI_{detail.cod}_{detail.sigla}_{comp or 'sem_competencia'}.xlsx"
    return StreamingResponse(
        io.BytesIO(excel_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── helpers ───────────────────────────────────────────────────────────────────

import re as _re
import decimal as _decimal
from datetime import datetime as _datetime, date as _date

_DT_RE   = _re.compile(r'^(\d{4})-(\d{2})-(\d{2})[T ](\d{2}):(\d{2}):(\d{2})')
_DATE_RE = _re.compile(r'^(\d{4})-(\d{2})-(\d{2})$')
_NUM_RE  = _re.compile(r'^-?\d+(\.\d+)?$')


def _cell_val(v: object) -> object:
    """Retorna o valor adequado para uma célula Excel.

    int / float / Decimal → número nativo (calculável, decimal exibido pelo locale do Excel)
    String puramente numérica → float ou int (cobre colunas varchar do TICKETING como tr, horas_em_pendente)
    String de data/datetime  → dd/mm/aa HH:MM:SS
    Qualquer outra string    → string
    """
    if v is None:
        return ""
    if isinstance(v, bool):
        return str(v)
    if isinstance(v, int):
        return v
    if isinstance(v, (float, _decimal.Decimal)):
        return float(v)
    s = str(v).strip()
    if _NUM_RE.match(s):
        try:
            return int(s) if "." not in s else float(s)
        except ValueError:
            pass
    m = _DT_RE.match(s)
    if m:
        y, mo, d, hh, mm, ss = m.groups()
        return f"{d}/{mo}/{y[2:]} {hh}:{mm}:{ss}"
    m = _DATE_RE.match(s)
    if m:
        y, mo, d = m.groups()
        return f"{d}/{mo}/{y[2:]} 00:00:00"
    return s


def _build_detail_excel(detail: KpiDetail, competencia: str) -> bytes:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    HEADER_FILL = PatternFill(start_color="205DF5", end_color="205DF5", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"KPI {int(detail.cod)} - {detail.sigla}"[:31]

    linhas = detail.linhas
    if not linhas:
        ws.append(["Sem dados para o período selecionado."])
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    headers = list(linhas[0].keys())

    # Cabeçalho
    ws.append(headers)
    ws.row_dimensions[1].height = 20
    for cell in ws[1]:
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Linhas de dados
    for row in linhas:
        ws.append([_cell_val(row.get(h)) for h in headers])

    # Largura automática das colunas
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 3, 45)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
