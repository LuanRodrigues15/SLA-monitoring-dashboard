"""
Builds the ZIP package matching the Auditor system's expected format.

Excel: "ListaIndicadores" sheet + 31 sheets named "01"–"31"
TXTs : 31 pipe-delimited files named "01 (MNT).txt", "02 (DMA).txt", …
Schema (14 columns):
  ID_arquivo | Codigo | Categoria | Detalhamento |
  Texto_01 | Texto_02 | Texto_03 | Texto_04 | Texto_05 |
  DT_inicial | DT_final | Medicao | HoraArquivo | Competencia

Granularidade das linhas de detalhe:
  - Zabbix Disp (02,04,05,07,08): 1 linha por evento de downtime
  - EBA (03)                    : 1 linha por site do segmento A (inventário: TIPOA + 10G + inventario de equipamentos)
  - EBC (06)                    : 1 linha por host (inventário)
  - MNT (01)                    : 1 linha por registro de manutenção
  - TICKETING (09-31)                 : 1 linha por ticket
"""
from __future__ import annotations

import calendar
import io
import re
import zipfile
from datetime import datetime, date as _date
from typing import Any

import pytz
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from app.config import settings
from app.core.trino_client import TrinoClient
from app.services.kpi_calculator import (
    KPI_CATALOG, _TICKETING_KPI_CONFIG, _ZABBIX_DISP_CONFIG, _ZABBIX_INV_CONFIG,
    _ops, _sz, _gz, _ticketing,
    dma_intervals_sql, _eba_03_sql,
    _has_consolidado,
)

_FUSO = pytz.timezone("America/Sao_Paulo")
_RE_HOST = re.compile(r'^([A-Za-z]+)(\d+)')


def _host_to_ref_externa(host: str) -> str:
    m = _RE_HOST.match(host)
    return (m.group(1).upper() + m.group(2).zfill(4)) if m else host


COLUNAS_FINAIS = [
    'ID_arquivo', 'Codigo', 'Categoria', 'Detalhamento',
    'Texto_01', 'Texto_02', 'Texto_03', 'Texto_04', 'Texto_05',
    'DT_inicial', 'DT_final', 'Medicao', 'HoraArquivo', 'Competencia',
]

MENSAGEM_SEM_DADOS = (
    "No período aferido, não houve dados ou chamados registrados para este indicador."
)

_LISTA_DESCR: dict[str, str] = {
    '01': '1. Manutenção Corretiva de Enlace - MNT',
    '02': '2. Disponibilidade Mensal — Enlace de Dados - DMA',
    '03': '3. Entrega de Banda — Enlace de Dados – EBA',
    '04': '4. Disponibilidade Mensal — Telefonia IP – DMB',
    '05': '5. Disponibilidade Mensal — Conectividade Sem Fio – DMC',
    '06': '6. Entrega de Banda — Conectividade Sem Fio – EBC',
    '07': '7. Disponibilidade Mensal — Monitoramento por Vídeo – DMD',
    '08': '8. Disponibilidade Mensal — Serviço Complementar – DME',
    '09': '9. Qualidade do Serviço de Operação – SO1',
    '10': '10. Qualidade de Satisfação no Atendimento – SO2',
    '11': '11. Grau de Satisfação — Telefonia IP – SO3',
    '12': '12. Tempo de Resposta — Enlace de Dados – TRA',
    '13': '13. Tempo de Solução — Enlace de Dados – TSA',
    '14': '14. Efetividade — Enlace de Dados – EFA',
    '15': '15. Reabertura — Enlace de Dados – RBA',
    '16': '16. Tempo de Resposta — Telefonia IP – TRB',
    '17': '17. Tempo de Solução — Telefonia IP – TSB',
    '18': '18. Efetividade — Telefonia IP – EFB',
    '19': '19. Reabertura — Telefonia IP – RBB',
    '20': '20. Tempo de Resposta — Conectividade Sem Fio – TRC',
    '21': '21. Tempo de Solução — Conectividade Sem Fio – TSC',
    '22': '22. Efetividade — Conectividade Sem Fio – EFC',
    '23': '23. Reabertura — Conectividade Sem Fio – RBC',
    '24': '24. Tempo de Resposta — Monitoramento por Vídeo – TRD',
    '25': '25. Tempo de Solução — Monitoramento por Vídeo – TSD',
    '26': '26. Efetividade — Monitoramento por Vídeo – EFD',
    '27': '27. Reabertura — Monitoramento por Vídeo – RBD',
    '28': '28. Tempo de Resposta — Serviço Complementar – TRE',
    '29': '29. Tempo de Solução — Serviço Complementar – TSE',
    '30': '30. Efetividade — Serviço Complementar – EFE',
    '31': '31. Reabertura — Serviço Complementar – RBE',
}


# ── Batch queries ──────────────────────────────────────────────────────────

def _batch_kpi_agg(trino: TrinoClient, competencia: str | None) -> dict[str, dict]:
    schema = _ops()
    if competencia:
        sql = (
            f"SELECT cod, kpi_valor, competencia FROM {schema}.kpi_agg_test"
            f" WHERE competencia = '{competencia}'"
        )
    else:
        sql = (
            f"WITH ranked AS ("
            f"  SELECT cod, kpi_valor, competencia,"
            f"         ROW_NUMBER() OVER (PARTITION BY cod ORDER BY competencia DESC, updated_at DESC) AS rn"
            f"  FROM {schema}.kpi_agg_test"
            f") SELECT cod, kpi_valor, competencia FROM ranked WHERE rn = 1"
            f" ORDER BY cod"
        )
    rows = trino.query_dict(sql)
    result: dict[str, dict] = {}
    for r in rows:
        cod = str(r["cod"])
        if cod not in result:
            result[cod] = r
    return result


def _batch_ticketing(trino: TrinoClient, comp: str | None) -> list[dict]:
    """Busca todos os tickets TICKETING do período, 1 linha por ticket (ROW_NUMBER dedup).

    Inclui tickets onde qualquer das 3 datas-chave cai no mês, pois cada tipo de KPI
    usa uma coluna diferente como filtro (TR→dt_triagem, demais→dt_resolvido_efetivo/dt_abertura).
    dt_resolvido_efetivo = COALESCE(dt_resolvido_real, dt_resolvido) — alinhado ao kpi_calculator.
    """
    h = _ticketing()
    mes_f = (
        f"AND (SUBSTR(dt_abertura, 1, 7) = '{comp}'"
        f" OR SUBSTR(dt_triagem, 1, 7) = '{comp}'"
        f" OR SUBSTR(COALESCE(dt_resolvido_real, dt_resolvido), 1, 7) = '{comp}')"
    ) if comp else ""
    try:
        return trino.query_dict(
            f"SELECT ticket_id, servico, dt_abertura, dt_triagem, dt_resolvido,"
            f" COALESCE(dt_resolvido_real, dt_resolvido) AS dt_resolvido_efetivo,"
            f" tr, ts, horas_em_pendente_cliente, houve_reabertura, qsos, qsat, qsri,"
            f" cliente_descricao, cod_ref_externa,"
            f" CASE WHEN dt_triagem IS NOT NULL AND dt_triagem != ''"
            f"           AND dt_abertura IS NOT NULL AND dt_abertura != '' THEN"
            f"   CAST(date_diff('second',"
            f"     TRY(date_parse(substr(dt_abertura,1,19),'%Y-%m-%dT%H:%i:%s')),"
            f"     TRY(date_parse(substr(dt_triagem,1,19),'%Y-%m-%dT%H:%i:%s'))"
            f"   ) AS DOUBLE) / 3600.0"
            f" ELSE NULL END AS tr_calculado,"
            f" CASE WHEN COALESCE(dt_resolvido_real, dt_resolvido) IS NOT NULL"
            f"           AND COALESCE(dt_resolvido_real, dt_resolvido) != '' THEN"
            f"   CAST(date_diff('second',"
            f"     TRY(date_parse(substr(dt_abertura,1,19),'%Y-%m-%dT%H:%i:%s')),"
            f"     TRY(date_parse(substr(COALESCE(dt_resolvido_real, dt_resolvido),1,19),'%Y-%m-%dT%H:%i:%s'))"
            f"   ) AS DOUBLE) / 3600.0"
            f"   - COALESCE(TRY_CAST(horas_em_pendente_cliente AS DOUBLE), 0.0)"
            f" ELSE NULL END AS ts_calculado"
            f" FROM ("
            f"   SELECT ticket_id, servico, dt_abertura, dt_triagem, dt_resolvido,"
            f"     dt_resolvido_real,"
            f"     tr, ts, horas_em_pendente_cliente, houve_reabertura, qsos, qsat, qsri,"
            f"     cliente_descricao, cod_ref_externa,"
            f"     ROW_NUMBER() OVER (PARTITION BY ticket_id ORDER BY COALESCE(dt_resolvido_real, dt_resolvido) DESC) AS rn"
            f"   FROM {h} WHERE TRUE {mes_f}"
            f" ) WHERE rn = 1"
            f" ORDER BY servico, dt_abertura DESC"
        )
    except Exception:
        return []


def _batch_mcfo(trino: TrinoClient, comp: str | None) -> list[dict]:
    schema = _ops()
    mes_f = f"AND competencia_referencia = '{comp}'" if comp else ""
    try:
        return trino.query_dict(
            f"SELECT trecho_origem, trecho_destino, dt_inicio, dt_final,"
            f" nc, extensao_km, m, ROUND(mcfo_valor,3) AS mcfo_valor, pontuacao, competencia_referencia"
            f" FROM {schema}.stage_mcfo_input WHERE deletado = false {mes_f} ORDER BY dt_final DESC"
        )
    except Exception:
        return []


def _batch_dma_02(trino: TrinoClient, comp: str | None) -> list[dict]:
    """KPI 02 — retorna intervalos de downtime via SQL de referência (billing + eventos de energia)."""
    if comp is None:
        now = datetime.now(_FUSO)
        comp = f"{now.year}-{now.month - 1:02d}" if now.month > 1 else f"{now.year - 1}-12"
    try:
        return trino.query_dict(dma_intervals_sql(comp))
    except Exception:
        return []


def _batch_zabbix_events(
    trino: TrinoClient,
    groupids: list[int],
    comp: str | None,
    trigger_include: list[str] | None = None,
    trigger_exclude: list[str] | None = None,
    min_duration_sec: int = 0,
) -> list[dict]:
    """Retorna 1 linha por intervalo de downtime (LEFT JOIN + MIN, lookback 1 mês)."""
    sz, gz = _sz(), _gz()
    groups = ", ".join(str(g) for g in groupids)

    trigger_clauses: list[str] = []
    if trigger_include:
        inc = " OR ".join(f"LOWER(gt.trigger_desc) LIKE '%{t}%'" for t in trigger_include)
        trigger_clauses.append(f"AND ({inc})")
    if trigger_exclude:
        for t in trigger_exclude:
            trigger_clauses.append(f"AND LOWER(gt.trigger_desc) NOT LIKE '%{t}%'")
    trigger_filter = " ".join(trigger_clauses)

    min_filter = (
        f"AND date_diff('second', data_inicial, COALESCE(data_final, TIMESTAMP '2100-01-01 00:00:00')) >= {min_duration_sec}"
        if min_duration_sec > 0 else ""
    )

    if comp:
        mes_cte = (
            f"SELECT DATE_TRUNC('month', DATE_PARSE('{comp}', '%Y-%m')) AS inicio,"
            f" DATE_TRUNC('month', DATE_PARSE('{comp}', '%Y-%m')) + INTERVAL '1' MONTH AS fim"
        )
    else:
        mes_cte = (
            f"SELECT DATE_TRUNC('month', MAX(event_datetime)) AS inicio,"
            f" DATE_TRUNC('month', MAX(event_datetime)) + INTERVAL '1' MONTH AS fim"
            f" FROM {gz}.gold_events_triggers"
        )

    try:
        return trino.query_dict(f"""
WITH
mes AS ({mes_cte}),
hosts AS (
    SELECT DISTINCT h.hostid, h.host
    FROM {sz}.hosts h
    JOIN {sz}.hosts_groups hg ON hg.hostid = h.hostid
    WHERE hg.groupid IN ({groups}) AND h.status = 0
),
eventos_brutos AS (
    SELECT gt.host_id, gt.trigger_id, gt.trigger_desc, gt.event_datetime, gt.event_value
    FROM {gz}.gold_events_triggers gt
    JOIN hosts h ON h.hostid = gt.host_id
    JOIN mes m ON gt.event_datetime >= m.inicio - INTERVAL '1' MONTH
             AND gt.event_datetime <  m.fim + INTERVAL '7' DAY
    WHERE TRUE {trigger_filter}
),
intervalos AS (
    SELECT
        t1.host_id, t1.trigger_id, t1.trigger_desc,
        t1.event_datetime          AS data_inicial,
        MIN(t2.event_datetime)     AS data_final
    FROM eventos_brutos t1
    LEFT JOIN eventos_brutos t2
        ON  t1.host_id    = t2.host_id
        AND t1.trigger_id = t2.trigger_id
        AND t2.event_value = 0
        AND t2.event_datetime > t1.event_datetime
    WHERE t1.event_value = 1
    GROUP BY t1.host_id, t1.trigger_id, t1.trigger_desc, t1.event_datetime
),
intervalos_no_mes AS (
    SELECT i.host_id, i.trigger_desc, i.data_inicial,
           COALESCE(i.data_final, TIMESTAMP '2100-01-01 00:00:00') AS data_final
    FROM intervalos i
    CROSS JOIN mes m
    WHERE i.data_inicial < m.fim
      AND COALESCE(i.data_final, TIMESTAMP '2100-01-01 00:00:00') >= m.inicio
      {min_filter}
)
SELECT h.host,
       i.trigger_desc,
       GREATEST(i.data_inicial, m.inicio)                        AS dt_inicio_evento,
       LEAST(i.data_final, m.fim - INTERVAL '1' SECOND)         AS dt_fim_evento,
       date_diff('second',
           GREATEST(i.data_inicial, m.inicio),
           LEAST(i.data_final, m.fim - INTERVAL '1' SECOND)
       ) / 3600.0                                                AS horas_indisponivel
FROM intervalos_no_mes i
JOIN hosts h ON h.hostid = i.host_id
CROSS JOIN mes m
ORDER BY dt_inicio_evento, h.host
""")
    except Exception:
        return []


def _batch_zabbix_inv(trino: TrinoClient, cod: str) -> list[dict]:
    """Retorna 1 linha por host (inventário de banda comprometida). Usado para EBC (06)."""
    cfg = _ZABBIX_INV_CONFIG[cod]
    sz = _sz()
    groups = ", ".join(str(g) for g in cfg["groupids"])
    try:
        return trino.query_dict(
            f"SELECT h.host, {cfg['mbps']} AS mbps_comprometido"
            f" FROM {sz}.hosts h JOIN {sz}.hosts_groups hg ON hg.hostid = h.hostid"
            f" WHERE hg.groupid IN ({groups}) AND h.status = 0 ORDER BY h.host"
        )
    except Exception:
        return []


def _batch_eba_03(trino: TrinoClient) -> list[dict]:
    """Retorna inventário completo de sites do segmento A: TIPOA (grp201) + 10G (grp91) + inventario de equipamentos."""
    try:
        return trino.query_dict(_eba_03_sql())
    except Exception:
        return []


def _batch_consolidado(trino: TrinoClient, comp: str) -> dict[str, list[dict]]:
    """Lê linhas DETALHE de sla_meses_consolidado agrupadas por cod.

    Retorna dict cod → lista de dicts com as 14 colunas no formato COLUNAS_FINAIS,
    prontas para ser passadas diretamente ao construtor de TXT/Excel sem transformação.
    """
    schema = _ops()
    try:
        rows = trino.query_dict(
            f"SELECT cod, id_arquivo, categoria, detalhamento,"
            f" texto_01, texto_02, texto_03, texto_04, texto_05,"
            f" dt_inicial, dt_final, medicao, hora_arquivo"
            f" FROM {schema}.sla_meses_consolidado"
            f" WHERE competencia = '{comp}' AND tipo = 'DETALHE'"
            f" ORDER BY cod, dt_inicial"
        )
    except Exception:
        return {}

    by_cod: dict[str, list[dict]] = {}
    for r in rows:
        cod = str(r.get("cod", ""))
        by_cod.setdefault(cod, []).append({
            "ID_arquivo": str(r.get("id_arquivo") or "01"),
            "Categoria":  str(r.get("categoria") or ""),
            "Detalhamento": str(r.get("detalhamento") or ""),
            "Texto_01": _try_decimal_str(r.get("texto_01")),
            "Texto_02": _try_decimal_str(r.get("texto_02")),
            "Texto_03": _try_decimal_str(r.get("texto_03")),
            "Texto_04": _try_decimal_str(r.get("texto_04")),
            "Texto_05": _try_decimal_str(r.get("texto_05")),
            "DT_inicial": _fmt_dt(r.get("dt_inicial")),
            "DT_final":   _fmt_dt(r.get("dt_final")),
            "Medicao":    _try_numeric(r.get("medicao")),
            "HoraArquivo": _fmt_dt(r.get("hora_arquivo")),
        })
    return by_cod


# ── Helpers de schema Auditor ───────────────────────────────────────────────────

def _comp_info(comp: str | None) -> tuple[str, str, str, str]:
    """Returns (dt_inicial, dt_final, competencia_str, hora_arquivo)."""
    now = datetime.now(_FUSO)
    if comp:
        year, month = int(comp[:4]), int(comp[5:7])
    else:
        if now.month == 1:
            year, month = now.year - 1, 12
        else:
            year, month = now.year, now.month - 1
    last_day = calendar.monthrange(year, month)[1]
    dt_ini = f"01/{month:02d}/{year} 00:00:00"
    dt_fim = f"{last_day:02d}/{month:02d}/{year} 23:59:59"
    comp_str = f"{month:02d}/{year}"
    hora_str = now.strftime("%d/%m/%Y %H:%M:%S")
    return dt_ini, dt_fim, comp_str, hora_str


_TZ_OFFSET_RE = re.compile(r'([+-])(\d{2}):(\d{2})$')


def _try_numeric(val: Any) -> Any:
    """Converte qualquer string numérica para float (inclui inteiros como '5')."""
    if isinstance(val, (int, float)):
        return val
    if val is None or str(val).strip() == '':
        return ''
    try:
        return float(str(val).replace(',', '.'))
    except (ValueError, TypeError):
        return str(val) if val is not None else ''


def _try_decimal_str(val: Any) -> Any:
    """Converte para float somente strings com ponto ou vírgula decimal ('0.5'→0.5, '0,5'→0.5).
    Strings sem separador decimal ('5', 'Excelente', 'PAG001') ficam como string."""
    if isinstance(val, (int, float)):
        return val
    s = str(val).strip() if val is not None else ''
    if not s or ('.' not in s and ',' not in s):
        return s
    try:
        return float(s.replace(',', '.'))
    except (ValueError, TypeError):
        return s


def _fmt_dt(val: Any) -> str:
    """Converts any date/datetime value to dd/MM/yyyy HH:mm:ss."""
    if val is None or val == "":
        return ""
    try:
        if isinstance(val, _date) and not isinstance(val, datetime):
            return val.strftime("%d/%m/%Y 00:00:00")
        if isinstance(val, datetime):
            dt = val
        else:
            s = str(val).strip().replace('T', ' ')
            m = _TZ_OFFSET_RE.search(s)
            if m:
                from datetime import timezone, timedelta as _td
                sign = 1 if m.group(1) == '+' else -1
                offset = _td(hours=int(m.group(2)) * sign, minutes=int(m.group(3)) * sign)
                dt_str = s[:m.start()].strip()
                dt = datetime.fromisoformat(dt_str).replace(tzinfo=timezone(offset))
            else:
                dt = datetime.fromisoformat(s)
        if dt.tzinfo is None:
            dt = _FUSO.localize(dt)
        else:
            dt = dt.astimezone(_FUSO)
        return dt.strftime("%d/%m/%Y %H:%M:%S")
    except Exception:
        return str(val)


def _base_row(cod: str, descr: str, dt_ini: str, dt_fim: str,
              hora: str, comp: str) -> dict:
    return {
        'ID_arquivo': '02',
        'Codigo': cod,
        'Categoria': descr,
        'Detalhamento': '',
        'Texto_01': '',
        'Texto_02': '',
        'Texto_03': '',
        'Texto_04': '',
        'Texto_05': '',
        'DT_inicial': dt_ini,
        'DT_final': dt_fim,
        'Medicao': '',
        'HoraArquivo': hora,
        'Competencia': comp,
    }


def _no_data_rows(cod: str, descr: str, dt_ini: str, dt_fim: str,
                  hora: str, comp: str) -> list[dict]:
    row = _base_row(cod, descr, dt_ini, dt_fim, hora, comp)
    row['Texto_01'] = MENSAGEM_SEM_DADOS
    row['DT_inicial'] = ''
    row['DT_final'] = ''
    return [row]


def _apply_id_arquivo(rows: list[dict]) -> list[dict]:
    """Última linha recebe ID_arquivo='02', demais '01' (padrão Auditor)."""
    for i, r in enumerate(rows):
        r['ID_arquivo'] = '02' if i == len(rows) - 1 else '01'
    return rows


def _sort_rows_by_dt_inicial(rows: list[dict]) -> list[dict]:
    """Ordena linhas por DT_inicial crescente; linhas sem data ficam no fim."""
    if len(rows) <= 1:
        return rows

    def sort_key(item: tuple[int, dict]) -> tuple[int, datetime, int]:
        idx, row = item
        raw = str(row.get('DT_inicial') or '').strip()
        try:
            return (0, datetime.strptime(raw, "%d/%m/%Y %H:%M:%S"), idx)
        except ValueError:
            return (1, datetime.max, idx)

    return [row for _, row in sorted(enumerate(rows), key=sort_key)]


# ── Helpers para Medicao por tipo de KPI TICKETING ─────────────────────────────

def _ticketing_medicao(ticket: dict, cfg: dict) -> Any:
    tipo, col = cfg["tipo"], cfg["col"]
    try:
        if tipo == "sat":
            return ticket.get(col, "") or ""
        elif tipo == "tr":
            v = ticket.get("tr_calculado")
            return round(float(v), 4) if v is not None else ""
        elif tipo == "ts":
            v = ticket.get("ts_calculado")
            return round(float(v), 4) if v is not None else ""
    except (ValueError, TypeError):
        pass
    return ""


def _ticketing_dt_final(ticket: dict, tipo: str) -> str:
    if tipo == "tr":
        return _fmt_dt(ticket.get("dt_triagem", ""))
    return _fmt_dt(ticket.get("dt_resolvido_efetivo", ""))


# ── Construtor de linhas por indicador ────────────────────────────────────

def _build_rows(cod: str, kpi: dict, valor: float | None, linhas: list[dict],
                dt_ini: str, dt_fim: str, hora: str, comp: str) -> list[dict]:
    descr = _LISTA_DESCR.get(cod, kpi.get('nome', ''))

    # MNT (01) — 1 linha por registro de manutenção
    if cod == '01':
        if not linhas:
            return _no_data_rows(cod, descr, dt_ini, dt_fim, hora, comp)
        rows = []
        for r in linhas:
            row = _base_row(cod, descr, dt_ini, dt_fim, hora, comp)
            row['Detalhamento'] = f"{r.get('trecho_origem', '')} ↔ {r.get('trecho_destino', '')}"
            row['Texto_01'] = str(r.get('nc', ''))
            row['Texto_02'] = r.get('extensao_km') or ''
            row['DT_inicial'] = _fmt_dt(r.get('dt_inicio'))
            row['DT_final'] = _fmt_dt(r.get('dt_final'))
            rows.append(row)
        return _apply_id_arquivo(rows)

    # KPI 02 (DMA)
    if cod == '02':
        if not linhas:
            return _no_data_rows(cod, descr, dt_ini, dt_fim, hora, comp)
        rows = []
        for r in linhas:
            row = _base_row(cod, descr, dt_ini, dt_fim, hora, comp)
            row['ID_arquivo'] = str(r.get('ID_arquivo', '02'))
            row['Detalhamento'] = str(r.get('Detalhamento', ''))
            row['Texto_01'] = str(r.get('Texto_01', '') or '')
            row['Texto_02'] = str(r.get('Texto_02', '') or '')
            row['Texto_03'] = str(r.get('Texto_03', '') or '')
            row['Texto_04'] = str(r.get('Texto_04', '') or '')
            row['Texto_05'] = str(r.get('Texto_05', '') or '')
            row['DT_inicial'] = _fmt_dt(r.get('DT_inicial'))
            row['DT_final'] = _fmt_dt(r.get('DT_final'))
            row['Medicao'] = r.get('horas_indisponivel', '')
            rows.append(row)
        return rows

    # EBA (03) — inventário: 1 linha por site do segmento A com Medicao fixo '30 Mbps'
    if cod == '03':
        if not linhas:
            return _no_data_rows(cod, descr, dt_ini, dt_fim, hora, comp)
        rows = []
        for r in linhas:
            host = str(r.get('host', ''))
            row = _base_row(cod, descr, dt_ini, dt_fim, hora, comp)
            row['Detalhamento'] = host
            row['Texto_01'] = host.split('_')[0]
            row['Texto_05'] = str(r.get('Ref_Externa', '') or '')
            row['DT_inicial'] = ''
            row['DT_final'] = ''
            row['Medicao'] = '30 Mbps'
            rows.append(row)
        return _apply_id_arquivo(rows)

    # Zabbix disponibilidade (04, 05, 07, 08) — 1 linha por intervalo de downtime
    if cod in _ZABBIX_DISP_CONFIG:
        if not linhas:
            return _no_data_rows(cod, descr, dt_ini, dt_fim, hora, comp)
        rows = []
        for r in linhas:
            host = str(r.get('host', ''))
            row = _base_row(cod, descr, dt_ini, dt_fim, hora, comp)
            if cod == '04':
                row['Detalhamento'] = host
                row['Texto_01'] = str(r.get('trigger_desc', '') or '')
                row['Texto_02'] = ''
                row['Texto_03'] = ''
                row['Texto_04'] = ''
                row['Texto_05'] = ''
            else:
                row['Detalhamento'] = host.replace(' ', '_').split('_')[0]
                row['Texto_01'] = str(r.get('trigger_desc', '') or '')
                row['Texto_02'] = host
                row['Texto_04'] = ''
                row['Texto_05'] = _host_to_ref_externa(host)
            row['DT_inicial'] = _fmt_dt(r.get('dt_inicio_evento'))
            row['DT_final'] = _fmt_dt(r.get('dt_fim_evento'))
            row['Medicao'] = r.get('horas_indisponivel', '')
            rows.append(row)
        return _apply_id_arquivo(rows)

    # EBC (06) — inventário: 1 linha por host AP com Medicao fixo '45 Mbps'
    if cod == '06':
        if not linhas:
            return _no_data_rows(cod, descr, dt_ini, dt_fim, hora, comp)
        rows = []
        for r in linhas:
            host = str(r.get('host', ''))
            row = _base_row(cod, descr, dt_ini, dt_fim, hora, comp)
            row['Detalhamento'] = host
            row['Texto_01'] = host.replace(' ', '_').split('_')[0]
            row['Texto_05'] = _host_to_ref_externa(host)
            row['DT_inicial'] = ''
            row['DT_final'] = ''
            row['Medicao'] = '45 Mbps'
            rows.append(row)
        return _apply_id_arquivo(rows)

    # TICKETING KPIs (09–31)
    if cod in _TICKETING_KPI_CONFIG:
        cfg = _TICKETING_KPI_CONFIG[cod]
        tipo = cfg['tipo']

        if not linhas:
            return _no_data_rows(cod, descr, dt_ini, dt_fim, hora, comp)

        # EA: 1 linha agregada
        # Texto_01 = TF (resolvidos no mês, numerador)  — formato sistema legado
        # Texto_02 = TA (abertos no mês, denominador)   — formato sistema legado
        # EA = TF/TA * 100; pode ultrapassar 100% quando backlog é zerado
        if tipo == 'ea':
            month_str = f"{comp[-4:]}-{comp[:2]}"
            tf = sum(1 for t in linhas if t.get('dt_resolvido_efetivo', '')[:7] == month_str)
            ta = sum(1 for t in linhas if t.get('dt_abertura', '')[:7] == month_str)
            row = _base_row(cod, descr, dt_ini, dt_fim, hora, comp)
            row['Texto_01'] = str(tf)   # resolvidos no mês (numerador)
            row['Texto_02'] = str(ta)   # abertos no mês (denominador)
            row['DT_inicial'] = ''
            row['DT_final'] = ''
            return [row]

        # RT: 1 linha agregada — reaberturas (Texto_01) e total resolvidos no mês (Texto_02)
        if tipo == 'rt':
            reopened = sum(1 for t in linhas if t.get('houve_reabertura') == 'Sim')
            row = _base_row(cod, descr, dt_ini, dt_fim, hora, comp)
            row['Texto_01'] = str(reopened)
            row['Texto_02'] = str(len(linhas))
            row['DT_inicial'] = ''
            row['DT_final'] = ''
            return [row]

        # sat: somente tickets com avaliação preenchida (alinhado ao comportamento do sistema legado)
        if tipo == 'sat':
            col_sat = cfg['col']
            linhas = [t for t in linhas if t.get(col_sat) not in (None, '')]
            if not linhas:
                return _no_data_rows(cod, descr, dt_ini, dt_fim, hora, comp)

        # sat, tr, ts: 1 linha por chamado
        rows = []
        for t in linhas:
            row = _base_row(cod, descr, dt_ini, dt_fim, hora, comp)
            row['Detalhamento'] = str(t.get('ticket_id', ''))
            row['Texto_01'] = str(t.get('cliente_descricao', '') or '')
            if tipo == 'ts':
                _pend = t.get('horas_em_pendente_cliente')
                if _pend is not None and _pend != '':
                    try:
                        row['Texto_02'] = float(_pend)
                    except (ValueError, TypeError):
                        row['Texto_02'] = str(_pend)
                else:
                    row['Texto_02'] = ''
            row['Texto_05'] = str(t.get('cod_ref_externa', '') or '')
            row['DT_inicial'] = _fmt_dt(t.get('dt_abertura', ''))
            row['DT_final'] = _ticketing_dt_final(t, tipo)
            row['Medicao'] = _ticketing_medicao(t, cfg)
            rows.append(row)
        return _apply_id_arquivo(rows)

    return _no_data_rows(cod, descr, dt_ini, dt_fim, hora, comp)


# ── Formatação Excel ───────────────────────────────────────────────────────

def _ajustar_planilha(wb: openpyxl.Workbook) -> None:
    grey_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    beige_fill = PatternFill(start_color="DDD9C4", end_color="DDD9C4", fill_type="solid")
    header_font = Font(bold=True)
    NUMBER_FORMAT = "#,##0.000"

    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = True
        header = [cell.value for cell in ws[1]]

        fill_header = None
        if ws.title == "ListaIndicadores":
            fill_header = grey_fill
        elif ws.title in ("03", "06"):
            fill_header = beige_fill

        for cell in ws[1]:
            cell.font = header_font
            cell.alignment = Alignment(wrap_text=False, vertical="top")
            if fill_header:
                cell.fill = fill_header

        id_col = next((i + 1 for i, n in enumerate(header) if n == "ID_arquivo"), -1)
        med_col = next((i + 1 for i, n in enumerate(header) if n == "Medicao"), -1)

        for r in range(1, ws.max_row + 1):
            ws.row_dimensions[r].height = 18
            for cell in ws[r]:
                cell.alignment = Alignment(wrap_text=False, vertical="top")
            if r > 1:
                if id_col != -1 and str(ws.cell(row=r, column=id_col).value) == "02":
                    ws.cell(row=r, column=id_col).fill = yellow_fill
                if med_col != -1:
                    m = ws.cell(row=r, column=med_col)
                    if isinstance(m.value, (int, float)):
                        m.number_format = NUMBER_FORMAT

        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=0)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 2


# ── TXT ────────────────────────────────────────────────────────────────────

def _rows_to_txt(rows: list[dict]) -> bytes:
    lines: list[str] = []
    for row in rows:
        parts: list[str] = []
        for col in COLUNAS_FINAIS:
            v = row[col]
            if isinstance(v, float):
                parts.append(str(v).replace(".", ","))
            elif v is None:
                parts.append("")
            else:
                s = str(v)
                if "." in s:
                    try:
                        float(s)
                        s = s.replace(".", ",")
                    except ValueError:
                        pass
                parts.append(s)
        lines.append("|".join(parts))
    return ("\r\n".join(lines) + "\r\n").encode("utf-8")


# ── Construção do ZIP ──────────────────────────────────────────────────────

def build_zip(
    trino: TrinoClient,
    competencia: str | None = None,
    content_type: str = "full",
) -> bytes:
    """Gera o pacote de indicadores.

    content_type:
        'full'  -> ZIP com Excel + 31 TXTs (padrão)
        'excel' -> bytes brutos do .xlsx
        'txts'  -> ZIP contendo apenas os 31 TXTs
    """
    # 1. Busca KPI values em batch
    kpi_vals = _batch_kpi_agg(trino, competencia)
    comps_usadas = {v.get("competencia") for v in kpi_vals.values() if v.get("competencia")}
    comp_detalhe = competencia or (max(comps_usadas) if comps_usadas else None)

    dt_ini, dt_fim, comp_str, hora_str = _comp_info(comp_detalhe)
    now = datetime.now(_FUSO)

    # 2. Busca dados de detalhe em batch
    # Para meses fechados/históricos: lê snapshot de sla_meses_consolidado (zero recálculo).
    # Para mês corrente aberto: consulta fontes ao vivo (TICKETING, Zabbix, stage_mcfo_input).
    is_consolidado = bool(comp_detalhe and _has_consolidado(trino, comp_detalhe))

    if is_consolidado:
        consolidado_by_cod = _batch_consolidado(trino, comp_detalhe)
        all_ticketing = []
        mcfo_rows = []
        inv_rows_c = []
        dma_02_rows = []
        eba_03_rows = []
        zabbix_events = {}
    else:
        consolidado_by_cod = {}
        all_ticketing = _batch_ticketing(trino, comp_detalhe)
        mcfo_rows = _batch_mcfo(trino, comp_detalhe)
        inv_rows_c = _batch_zabbix_inv(trino, "06")

        # KPI 02 (DMA): SQL de referência completo (billing + eventos de energia)
        dma_02_rows = _batch_dma_02(trino, comp_detalhe)

        # KPI 03 (EBA): inventário completo PAG — TIPOA (grp201) + 10G (grp91) + inventario de equipamentos
        eba_03_rows = _batch_eba_03(trino)

        # Eventos de downtime: KPIs Disp exceto 02 (que tem lógica própria) e 03 (inventário)
        zabbix_events: dict[str, list[dict]] = {}
        for cod, cfg in _ZABBIX_DISP_CONFIG.items():
            if cod == "02":
                continue  # tratado separadamente via dma_02_rows
            zabbix_events[cod] = _batch_zabbix_events(
                trino, cfg["groupids"], comp_detalhe,
                cfg["trigger_include"], cfg["trigger_exclude"],
                cfg.get("min_duration_sec", 0),
            )

    # 3. Monta Excel
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws_lista = wb.create_sheet("ListaIndicadores")
    ws_lista.append(["Codigo", "Indicador", "Sigla"])
    for kpi in KPI_CATALOG:
        ws_lista.append([kpi["cod"], _LISTA_DESCR.get(kpi["cod"], kpi["nome"]), kpi["sigla"]])

    all_txts: dict[str, bytes] = {}

    for kpi in KPI_CATALOG:
        cod = kpi["cod"]
        sigla = kpi["sigla"]
        row_data = kpi_vals.get(cod, {})
        valor = float(row_data["kpi_valor"]) if row_data.get("kpi_valor") is not None else None

        # Seleciona as linhas de detalhe corretas
        if is_consolidado:
            # Mês fechado/histórico: reconstituição direta do snapshot
            raw = consolidado_by_cod.get(cod, [])
            if raw:
                descr = _LISTA_DESCR.get(cod, kpi.get("nome", ""))
                rows = []
                for r in raw:
                    rows.append({
                        "ID_arquivo":   r["ID_arquivo"],
                        "Codigo":       cod,
                        "Categoria":    descr,
                        "Detalhamento": r["Detalhamento"],
                        "Texto_01":     r["Texto_01"],
                        "Texto_02":     r["Texto_02"],
                        "Texto_03":     r["Texto_03"],
                        "Texto_04":     r["Texto_04"],
                        "Texto_05":     r["Texto_05"],
                        "DT_inicial":   r["DT_inicial"],
                        "DT_final":     r["DT_final"],
                        "Medicao":      r["Medicao"],
                        "HoraArquivo":  r["HoraArquivo"] or hora_str,
                        "Competencia":  comp_str,
                    })
            else:
                rows = _no_data_rows(
                    cod, _LISTA_DESCR.get(cod, kpi.get("nome", "")),
                    dt_ini, dt_fim, hora_str, comp_str,
                )
        else:
            # Mês aberto: fontes ao vivo
            if cod == "01":
                linhas = mcfo_rows
            elif cod == "02":
                linhas = dma_02_rows
            elif cod == "03":
                linhas = eba_03_rows
            elif cod == "06":
                linhas = inv_rows_c
            elif cod in zabbix_events:  # 04, 05, 07, 08
                linhas = zabbix_events[cod]
            elif cod in _TICKETING_KPI_CONFIG:
                cfg_h = _TICKETING_KPI_CONFIG[cod]
                svc = cfg_h["servico"]
                tipo_h = cfg_h["tipo"]
                base = [t for t in all_ticketing if svc is None or t.get("servico") == svc]
                if comp_detalhe:
                    if tipo_h == "tr":
                        linhas = [t for t in base if t.get("dt_triagem", "")[:7] == comp_detalhe]
                    elif tipo_h == "ea":
                        linhas = [t for t in base if t.get("dt_abertura", "")[:7] == comp_detalhe]
                    else:
                        linhas = [t for t in base if t.get("dt_abertura", "")[:7] == comp_detalhe]
                else:
                    linhas = base
            else:
                linhas = []

            rows = _build_rows(cod, kpi, valor, linhas, dt_ini, dt_fim, hora_str, comp_str)
            rows = _apply_id_arquivo(_sort_rows_by_dt_inicial(rows))

        # Aba Excel
        ws = wb.create_sheet(cod)
        ws.append(COLUNAS_FINAIS)
        for row in rows:
            ws.append([row[c] for c in COLUNAS_FINAIS])

        # TXT
        all_txts[f"{cod} ({sigla}).txt"] = _rows_to_txt(rows)

    _ajustar_planilha(wb)

    excel_buf = io.BytesIO()
    wb.save(excel_buf)

    # 4. Monta saída conforme content_type
    if content_type == "excel":
        return excel_buf.getvalue()

    if comp_detalhe:
        nome_excel = f"Relatório Indicadores SLA Monitoring Dashboard {comp_detalhe[5:7]}-{comp_detalhe[:4]}.xlsx"
    else:
        nome_excel = now.strftime("Relatório Indicadores SLA Monitoring Dashboard %m-%Y.xlsx")
    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        if content_type != "txts":
            zf.writestr(nome_excel, excel_buf.getvalue())
        for name, content in all_txts.items():
            zf.writestr(name, content)

    return zip_buf.getvalue()
